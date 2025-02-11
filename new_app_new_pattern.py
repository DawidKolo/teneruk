import re
import os
import datetime
import sys
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# variables
path = "C:\\test\\000"
path_out = "C:\\test\\001"
filelist = os.listdir(path)
week_xlsx_name = "\\week.xlsx"
month_xlsx_name = "\\month.xlsx"
patterns = [r"PRECISELY" + "_" + "[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]" + ".PDF"]
for p in range(1,9):
    patterns.append(fr"P{p}" + "_" + "[0-9][0-9][0-9][0-9]" + "_" + "[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]" + ".PDF")

myFilename_month = path_out + month_xlsx_name
myFilename_week = path_out + week_xlsx_name


# checks if the xlsx file exists -> if not creates a new one
def check_xls_file():
    #    print("check_xls_file")
    if not os.path.isfile(path_out + month_xlsx_name):
        excel = xlsxwriter.Workbook(path_out + month_xlsx_name)
        excel.close()


# checks if the xlsx file exists -> if not creates a new one
def week_check_xls_file():
    #    print("week_check_xls_file")
    if not os.path.isfile(path_out + week_xlsx_name):
        excel = xlsxwriter.Workbook(path_out + week_xlsx_name)
        excel.close()


# gets filenames od txt files and adds them to the list, returns the list of strings
def collect_txt_filenames():
    #    print("collect_txt_filenames")
    txt_files = []  # creates a list
    for file in filelist:  # iterates through files in folder and adds their filenames ending with txt, to the list
        if file.endswith(".txt"):
            txt_files.append(file)

    return txt_files


def week_or_month_of_year(date, var):
    year = int(date[:4])
    month = int(date[5:6])
    day = int(date[6:9])
    if var == "w":
        week = datetime.date(year, month, day).isocalendar()[1]  # calculates the week of the year for a timestamp
        week = f"{week:02d}"
        week_in_year = (str(week) + "-" + str(year))  # adds week-year strings to the variable

        return week_in_year


def write_to_file(name, type_w_y, list):
    if type_w_y == "m":
        if not os.path.isfile(path_out + "\\" + name):
            y = open(path_out + "\\" + name, "w")
            for t in list:
                if int(name[7:9]) == int(t[0][4:6]):
                    line = ''.join(str(t))
                    y.write(line + "\n")
            y.close()
        else:
            y = open(path_out + "\\" + name, "w")
            for t in list:
                if int(name[7:9]) == int(t[0][4:6]):
                    line = ''.join(str(t))
                    y.write(line + "\n")
            y.close()

    elif type_w_y == "w":
        if not os.path.isfile(path_out + "\\" + name):
            y = open(path_out + "\\" + name, "w")
            for t in list:
                if int(name[8:10]) == int(t[1][0:2]):
                    line = ''.join(str(t))
                    y.write(line + "\n")
            y.close()
        else:
            y = open(path_out + "\\" + name, "w")
            for t in list:
                if int(name[8:10]) == int(t[1][0:2]):
                    line = ''.join(str(t))
                    y.write(line + "\n")
            y.close()


def undef_strings():  # this function writes unexpected values to a file
    # Patterns to exclude
    pattern = r"P[0-9]" + "_" + "[0-9][0-9][0-9][0-9]" + "_" + "[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]" + ".PDF"
    pattern2 = r"PRECISELY" + "_" + "[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]" + ".PDF"

    all_files = collect_txt_filenames()  # gets list od files

    test = []
    for fle in all_files:  # looping through the files and gets all the content, removes new line chars
        j = open(path + "\\" + fle, "r")
        for line in j:
            top = line.replace("\n", "")
            test.append((fle[20:28], week_or_month_of_year(fle[20:28], "w"), top))

    filtered = []
    for und_expr in test:  # stores unexpected entries
        if not re.match(pattern, und_expr[2].upper()) and not re.match(pattern2, und_expr[2].upper()):
            filtered.append(und_expr)
    # splits program into path in respect of the choice initially made [m / w]
    if type_of_report == "m":
        if len(filtered) > 0:
            unexpected_txt_filenames_stage_1 = []
            for r in range(len(filtered)):
                unexpected_txt_filenames_stage_1.append(filtered[r][0][0:6])
            unexpected_txt_filenames_stage_1 = list(dict.fromkeys(unexpected_txt_filenames_stage_1))

            unexpected_txt_filenames_stage_2 = []
            for q in unexpected_txt_filenames_stage_1:
                unexpected_txt_filenames_stage_2.append("uf_" + q[0:6] + ".txt")

            for e in range(len(unexpected_txt_filenames_stage_2)):
                un_exp_file = unexpected_txt_filenames_stage_2[e]
                write_to_file(un_exp_file, type_of_report, filtered)
        print("There are additional unexpected entries! Check files for details: ", end=" ")
        for c in unexpected_txt_filenames_stage_2:
            print(c, end=" ")

    elif type_of_report == "w":
        if len(filtered) > 0:
            w_unexpected_txt_file_s_1 = []
            for e in filtered:
                filename_week = week_or_month_of_year(e[0], type_of_report)

                w_unexpected_txt_file_s_1.append(filename_week)
            w_unexpected_txt_file_s_1 = list(dict.fromkeys(w_unexpected_txt_file_s_1))

            w_unexpected_txt_file_s_2 = []
            for b in w_unexpected_txt_file_s_1:
                w_unexpected_txt_file_s_2.append("week_uf_" + b + ".txt")

            for g in w_unexpected_txt_file_s_2:
                un_exp_file = g
                write_to_file(un_exp_file, type_of_report, filtered)

            print(f"There are additional unexpected entries! Check files for details: ", end=" ")
            for c in w_unexpected_txt_file_s_2:
                print(c, end=" ")


# gets dates from the filenames and adds them to the list, returns list of tuples of integers
def get_week_of_year():
    #    print("get_week_of_year")
    date_for_week = create_sheetnames()[1]  # takes timestamps
    week_of_year = []  # list to store week of the year
    for i in range(len(date_for_week)):  # iterates through timestamps and assigns values to variables
        year = int(date_for_week[i][:4])
        month = int(date_for_week[i][5:7])
        day = int(date_for_week[i][8:10])
        week = datetime.date(year, month, day).isocalendar()[1]  # calculates the week of the year for a timestamp
        week_of_year.append((week, year))  # adds week-year tuples to the list

    return week_of_year


# puts styles to the spreadsheet. Fonts, background, lines
def paint_worksheets(ws, line):
    #    print("paint_worksheets")

    if type_of_report == "w":  # checks what type of report needs to be generated
        last_line = insert_values_to_spreadsheet_weekly()[
            line]  # last_line - parameter to paint thick line in the spreadsheet
    elif type_of_report == "m":
        last_line = insert_values_to_spreadsheet()[line]

    dimensions = [('A', 10), ('J', 12), ('K', 11), ('M', 45)]
    for d in dimensions:
        ws.column_dimensions[d[0]].width = d[1]  # setting column dimension for given columns

    cell_set = ws['A1':'L26']  # sets values in the center of the given cell
    for cell in cell_set:
        for c in cell:
            c.alignment = Alignment(horizontal='center')

    thick = Side(border_style="thick", color="000000")  # sets a style for border

    for border_cell in ws.iter_cols(min_col=1, max_col=10, min_row=last_line + 1, max_row=last_line + 1):
        for brdc in border_cell:  # puts a thick line at the bottom of values
            brdc.border = Border(bottom=thick)

    for bold_cell in ws.iter_cols(min_col=2, max_col=10, min_row=1, max_row=1):
        for bc in bold_cell:  # paints a background for given cells and
            bc.font = Font(bold=True)  # changes font to bold
            bc.fill = PatternFill(start_color="6600CC", end_color="6600CC", fill_type="solid")

    for rows in ws.iter_cols(min_row=1, max_row=1, min_col=12, max_col=13):
        for cell in rows:  # paints a background for given cells and
            cell.font = Font(bold=True)  # changes font to bold
            cell.fill = PatternFill(start_color="6600CC", end_color="6600CC", fill_type="solid")

    for t in ws.iter_cols(min_row=24, max_row=25, min_col=2, max_col=10):  # changes font to bold
        for cell in t:
            cell.font = Font(bold=True)


# writes keys to the spreadsheet and paints a thick line at the last row
def keys(ws):
    #    print("keys")
    thick = Side(border_style="thick", color="000000")
    # keys
    keys = ["Keys", "P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8", "PRECISELY"]
    for key in range(len(keys)):
        ws.cell(column=12, row=key + 1, value=keys[key])

    legend = ["Returned undeliverable mail", "Returned mail for processing", "Complaints", "WEB Loyalty mail",
              "Claims packages", "Keys/ Returned Key", "Tenerity Recorded mail / Special Delivery / Signed for",
              "Personal Recorded mail / Special Delivery / Signed for", "Precisely - goneaway"]
    for k in range(len(legend)):
        ws.cell(column=13, row=k + 2, value=legend[k])
    for cols in ws.iter_cols(min_col=12, max_col=13, min_row=10, max_row=10):
        for h in cols:
            h.border = Border(bottom=thick)


# creates sheetnames out of filenames, returns a tuple of lists of strings
def create_sheetnames():
    #    print("create_sheetnames")
    txt_files = collect_txt_filenames()  # takes list of txt filenames
    sheet_names = []
    timestamp = []
    for txt_file in txt_files:  # extracts dates from filenames and inserts 2X"-" in
        x = re.findall("[0-9]", txt_file)
        x.insert(4, "-")
        x.insert(7, "-")

        date = "".join(x)
        sheet_name = date[0:7]
        sheet_names.append(sheet_name)  # writes sheetnames to the list
        timestamp.append(date)  # writes timestamps to the list

    return sheet_names, timestamp


# creates sheetnames out of filenames, returns a tuple of lists of strings (for weekly path)
def create_sheetnames_weekly():
    #    print("create_sheetnames_weekly")
    week = get_week_of_year()  # takes a list of week numbers of a year
    txt_files = collect_txt_filenames()  # takes list of txt filenames
    weekly_sheetnames = []

    timestamp = []
    for txt_file in txt_files:  # extracts dates from filenames and inserts 2X"-" in
        x = re.findall("[0-9]", txt_file)
        x.insert(4, "-")
        x.insert(7, "-")
        date = "".join(x)
        timestamp.append(date)

    for item in week:  # creates sheetnames based on the number of the week of the year
        weekly_sheetnames.append(str(item[0]) + "-" + str(item[1]))

    return weekly_sheetnames, timestamp


# searches for P1,P2 etc. from the pattern list in txt files, returns a list
def search_for_p():
    #    print("search_for_p")
    txt_files = collect_txt_filenames()
    temp_list = []

    for txt_file in txt_files:  # opens txt files
        f = open(path + "\\" + txt_file, "r")
        files = f.read().upper()

        p_table = []
        for p in patterns:  # searches for P values in the txt files and writes them to the list
            match_p = re.findall(p, files)
            p_table.append(len(match_p))
        temp_list.append(p_table)
        f.close()

    return temp_list


# writes sheetname to the workbook
def write_sheetname_to_wb():
    #    print("write_sheetname_to_wb")
    wb = load_workbook(filename=myFilename_month)
    sheetname = list(dict.fromkeys(create_sheetnames()[0]))  # gets unique items from a list
    for c in range(len(sheetname)):
        if not sheetname[c] in wb.sheetnames:  # creates sheetnames in the workbook if they don't exist
            wb.create_sheet(sheetname[c])

    wb.save(filename=myFilename_month)
    wb.close()


# writes weekly related sheetnames to workbook
def write_weekly_sheetname_to_wb():
    #    print("write_weekly_sheetname_to_wb")
    wb_week = load_workbook(filename=myFilename_week)
    sheetname = list(dict.fromkeys(create_sheetnames_weekly()[0]))  # gets unique item from a list

    for c in range(len(sheetname)):
        if not sheetname[c] in wb_week.sheetnames:  # creates sheetnames in the workbook if they don't exist
            wb_week.create_sheet(sheetname[c])

    wb_week.save(filename=myFilename_week)
    wb_week.close()


# Writes the P values to the spreadsheets and returns list of number of entries per spreadsheet
def insert_values_to_spreadsheet():
    #    print("insert_values_to_spreadsheet")
    wb = load_workbook(filename=myFilename_month)
    sheetname = list(dict.fromkeys(create_sheetnames()[0]))  # gets unique items from a list
    date = create_sheetnames()[1]
    top_bar = ["P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8", "PRECISELY"]
    p_table = search_for_p()

    outer_val = []  # the list of lengths of lists inner_val
    for w in range(len(sheetname)):  # iterates through the list and get list of sheetnames
        ws = wb[sheetname[w]]

        for n in range(len(top_bar)):  # writes header of "P"s into the spreadsheet
            ws.cell(column=2 + n, row=1, value=top_bar[n])

        inner_val = []  # the list of dates
        for item in range(len(date)):  # iterates through date list
            if date[item][0:7] == sheetname[w]:  # condition: to write correct values to the right spreadsheet
                ws.cell(column=1, row=ws.max_row + 1, value=date[item])  # writes values to cells

                for p in range(len(patterns)):  # writes values of "P"s to the spreadsheet
                    ws.cell(column=2 + p, row=ws.max_row, value=p_table[item][p])
                inner_val.append(date[item])  # increases the length of inner_val list

        outer_val.append(len(inner_val))
    wb.save(filename=myFilename_month)
    wb.close()

    return outer_val


# inserts weekly values to spreadsheet
def insert_values_to_spreadsheet_weekly():
    #    print("insert_values_to_spreadsheet_weekly")
    wb_week = load_workbook(filename=myFilename_week)
    sh_name = list(dict.fromkeys(create_sheetnames_weekly()[0]))  # gets unique items from a list
    date = create_sheetnames_weekly()[1]  # gets list of sheetnames [week-year]
    p_table = search_for_p()  # gets values of "P"s
    w_y = get_week_of_year()
    top_bar = ["P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8", "PRECISELY"]

    outer_val = []  # the list of lengths of lists inner_val
    for w in range(len(sh_name)):  # iterates through the list and get list of sheetnames
        ws = wb_week[sh_name[w]]

        for n in range(0, len(top_bar)):  # writes header of "P"s into the spreadsheet
            ws.cell(column=2 + n, row=1, value=top_bar[n])

        list_week_year = []
        inner_val = []  # the list of dates

        for item in range(len(w_y)):  # creates a list of  strings [week-year]
            list_week_year.append(str(w_y[item][0]) + "-" + str(w_y[item][1]))

            if list_week_year[item] == sh_name[w]:  # checks if a spreadsheet exists in the workbook
                ws.cell(column=1, row=ws.max_row + 1, value=date[item])  # writes dates to the spreadsheet
                for p in range(0, len(patterns)):  # writes values of "P"s to spreadsheet
                    ws.cell(column=2 + p, row=ws.max_row, value=p_table[item][p])

                inner_val.append(date[item])  # adds values to the list

        outer_val.append(len(inner_val))  # adds values to the list
        wb_week.save(filename=myFilename_week)
        wb_week.close()

    return outer_val


# writes a row of sums to relevant columns
def insert_monthly_sums_to_spreadsheet():
    #    print("insert_monthly_sums_to_spreadsheet")
    wb = load_workbook(filename=myFilename_month)
    sh_name = list(dict.fromkeys(create_sheetnames()[0]))  # gets unique items from a list

    for sh in range(len(sh_name)):  # iterates through sheetnames
        ws = wb[sh_name[sh]]
        paint_worksheets(ws, sh)  # styles the spreadsheets

        sums2 = ["=SUM(B1:B24)", "=SUM(C1:C24)", "=SUM(D1:D24)", "=SUM(E1:E24)", "=SUM(F1:F24)", "=SUM(G1:G24)",
                 "=SUM(H1:H24)", "=SUM(I1:I24)", "=SUM(J1:J24)", "=SUM(B25:J25)"]
        for s2 in range(len(sums2)):  # inserts sums of columns to the spreadsheet
            ws.cell(column=s2 + 2, row=25, value=sums2[s2])
        ws.cell(column=11, row=24, value="Monthly Sum")

        keys(ws)  # inserts keys

    wb.save(filename=myFilename_month)
    wb.close()


def insert_weekly_sums_to_spreadsheet():
    #    print("insert_weekly_sums_to_spreadsheet")
    wb_week = load_workbook(filename=myFilename_week)
    sh_name = list(dict.fromkeys(create_sheetnames_weekly()[0]))  # gets unique items from a list

    for sh in range(len(sh_name)):  # iterates through sheetnames
        ws = wb_week[sh_name[sh]]
        paint_worksheets(ws, sh)  # styles the spreadsheets

        sums = ["=SUM(B1:B8)", "=SUM(C1:C8)", "=SUM(D1:D8)", "=SUM(E1:E8)", "=SUM(F1:F8)", "=SUM(G1:G8)", "=SUM(H1:H8)",
                "=SUM(I1:I8)", "=SUM(J1:J8)", "=SUM(B10:J10)"]

        for sum in range(len(sums)):  # inserts sums of columns to the spreadsheet
            ws.cell(column=sum + 2, row=10, value=sums[sum])
        ws.cell(column=11, row=9, value="Weekly Sum")

        keys(ws)  # inserts keys

    wb_week.save(filename=myFilename_week)
    wb_week.close()


# starts the program
type_of_report = input('What type of report would you like to create? [m - monthly | w - weekly] ').lower()
if type_of_report == 'm':
    check_xls_file()
    write_sheetname_to_wb()
    insert_values_to_spreadsheet()
    insert_monthly_sums_to_spreadsheet()
    undef_strings()

elif type_of_report == 'w':
    week_check_xls_file()
    write_weekly_sheetname_to_wb()
    insert_values_to_spreadsheet_weekly()
    insert_weekly_sums_to_spreadsheet()
    undef_strings()

else:
    print('Only m or w was expected.')
    sys.exit()
