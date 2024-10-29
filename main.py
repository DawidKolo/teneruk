import re
import os
from fileinput import filename

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils import rows_from_range

# Define Input and output paths below
path = "C:\\test\\000"
path_out = "C:\\test\\001"

# creates list of files in the path directory
filelist = os.listdir(path)
if not os.path.isfile(path_out + "\\" + "a.xlsx"):

    excell = xlsxwriter.Workbook(path_out + "\\" + 'a.xlsx')
    excell.close()
else:
    pass
# empty list to store the txt files
txt_files = []

# loop to go through files in the directory
for file in filelist:
    # check if the file has a txt extension
    if file.endswith(".txt"):
        # adding txt files to the list
        txt_files.append(file)

        # loop to go through txt filenames in the list and create a sheetname [date]
        for txt_file in txt_files:
            x = re.findall("[0-9]", txt_file)
            x.insert(4, "-")
            x.insert(7, "-")

            date = "".join(x)
            sheet_name = date[0:7]

            # searching for pattern in the txt files and counting them
            f = open(path + "\\" + txt_file, "r")
            files = f.read()
            pattern = ["P1", "P2", "P3", "P4"]
            p_table = []
            for p in pattern:
                match_p = re.findall(p, files)
                p_table.append(len(match_p))

            f.close()
        # print(date, p_table)
        myFileName = path_out + "\\" + "a.xlsx"

        # load the workbook, and put the sheet into a variable
        wb = load_workbook(filename=myFileName, data_only=True)

        # creating a sheet in the workbook
        if not sheet_name in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        # max_row is a sheet function that gets the last row in a sheet.
        newRowLocation = ws.max_row + 1

        # write to the cell you want, specifying row and column, and value :-)

        ws.cell(column=2, row=1, value='P1')
        ws.cell(column=3, row=1, value='P2')
        ws.cell(column=4, row=1, value='P3')
        ws.cell(column=5, row=1, value='P4')

        ws.cell(column=1, row=newRowLocation, value=date)
        ws.cell(column=2, row=newRowLocation, value=p_table[0])
        ws.cell(column=3, row=newRowLocation, value=p_table[1])
        ws.cell(column=4, row=newRowLocation, value=p_table[2])
        ws.cell(column=5, row=newRowLocation, value=p_table[3])

        wb.save(filename=myFileName)
        wb.close()

myFileName = path_out + "\\" + "a.xlsx"
sum_wb = load_workbook(filename=myFileName, data_only=True)
list_of_sheets = sum_wb.sheetnames

for sheet in list_of_sheets:
    # if sheet == 'Sheet1':
    #     del sum_wb['Sheet1']
    worksheet = sum_wb[sheet]
    rows_count = worksheet.max_row
    column_count = worksheet.max_column

    # Partial Sum and Total Monthly Sum
    first_row = 2
    last_row = rows_count
    sum_row = last_row + 3
    start_col = 2
    end_col = column_count

    for row in worksheet.iter_rows(min_row=sum_row, max_row=sum_row, min_col=start_col, max_col=end_col):
        cell_address = []

        for cell in row:
            cell_sum_start = cell.column_letter + str(first_row)
            cell_sum_end = cell.column_letter + str(last_row)
            cell.value = '=SUM({0}:{1})'.format(cell_sum_start, cell_sum_end)
            cell_address.append(('F', f'{cell.row}'))

        cell_address = list(dict.fromkeys(cell_address))

        for s in cell_address:
            a = s[1]
            z = int(a)

            B = worksheet[f'B{s[1]}'].value
            C = worksheet[f'C{s[1]}'].value
            cc = C.replace('=', '+', 1)

            D = worksheet[f'D{s[1]}'].value
            dd = D.replace('=', '+', 1)
            E = worksheet[f'E{s[1]}'].value
            ee = E.replace('=', '+', 1)

            worksheet['F'f'{s[1]}'].value = B + cc + dd + ee
            worksheet['F'f'{z - 1}'].value = "Monthly Sum"

sum_wb.save(filename=myFileName)
sum_wb.close()

# moving files from in directory to out directory
for x in filelist:
    os.rename(path + "\\" + x, path_out + "\\" + x)
