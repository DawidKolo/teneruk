import re
import os
import datetime
import sys

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
# variables
path = "C:\\test\\000"
path_out = "C:\\test\\001"

filelist = os.listdir(path)
week_xlsx_name = "\\week.xlsx"
month_xlsx_name = "\\month.xlsx"
pattern = ["P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8"]


# checks if the xlsx file exists -> if not creates a new one
def check_xls_file():
    if not os.path.isfile(path_out + month_xlsx_name):
        excel = xlsxwriter.Workbook(path_out + month_xlsx_name)
        excel.close()


def week_check_xls_file():
    if not os.path.isfile(path_out + week_xlsx_name):
        excel = xlsxwriter.Workbook(path_out + week_xlsx_name)
        excel.close()


# gets filenames od txt files and adds them to the list, returns the list of strings
def collect_txt_filenames():
    txt_files = []
    for file in filelist:
        if file.endswith(".txt"):
            txt_files.append(file)
    return txt_files


# gets dates from the filenames and adds them to the list, returns list of tuples of integers
def get_week_of_year():
    date_for_week = create_sheetnames()[1]
    week_of_year = []
    for i in range(len(date_for_week)):
        year = int(date_for_week[i][:4])
        month = int(date_for_week[i][5:7])
        day = int(date_for_week[i][8:10])
        week = datetime.date(year, month, day).isocalendar()[1]
        week_of_year.append((week, year))

    return week_of_year


def paint_worksheets(ws):

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['M'].width = 45
    cell_set = ws['A1':'L30']
    for cell in cell_set:
        for c in cell:
            c.alignment = Alignment(horizontal='center')

def keys(ws):
    # keys
    keys = ["Keys", "P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8"]
    for key in range(len(keys)):
        ws.cell(column=12, row=key + 2, value=keys[key])

    legend = ["Returned undeliverable mail", "Returned mail for processing", "Complaints", "WEB Loyalty mail",
              "Claims packages", "Keys/ Returned Key", "Tenerity Recorded mail / Special Delivery / Signed for",
              "Personal Recorded mail / Special Delivery / Signed for"]
    for k in range(len(legend)):
        ws.cell(column=13, row=k + 3, value=legend[k])




# creates a sheetnames out of filenames, returns a tuple of lists of strings
def create_sheetnames():
    txt_files = collect_txt_filenames()
    sheet_names = []
    timestamp = []
    for txt_file in txt_files:
        x = re.findall("[0-9]", txt_file)
        x.insert(4, "-")
        x.insert(7, "-")

        date = "".join(x)
        sheet_name = date[0:7]
        sheet_names.append(sheet_name)
        timestamp.append(date)

    return sheet_names, timestamp


def create_sheetnames_weekly():
    week = get_week_of_year()
    txt_files = collect_txt_filenames()
    weekly_sheetnames = []

    timestamp = []
    for txt_file in txt_files:
        x = re.findall("[0-9]", txt_file)
        x.insert(4, "-")
        x.insert(7, "-")
        date = "".join(x)
        timestamp.append(date)

    for item in week:
        weekly_sheetnames.append(str(item[0]) + "-" + str(item[1]))

    return weekly_sheetnames, timestamp


# searches for P1,P2 etc. from the pattern list in txt files, returns a list
def search_for_p():
    txt_files = collect_txt_filenames()
    temp_list = []
    for txt_file in txt_files:
        f = open(path + "\\" + txt_file, "r")
        files = f.read()

        p_table = []

        for p in pattern:
            match_p = re.findall(p, files)
            p_table.append(len(match_p))
        temp_list.append(p_table)
        f.close()

    return temp_list


# writes sheetname to the workbook
def write_sheetname_to_wb():
    sheetname = list(dict.fromkeys(create_sheetnames()[0]))
    for c in range(len(sheetname)):
        myFilename = path_out + month_xlsx_name

        wb = load_workbook(filename=myFilename)

        if not sheetname in wb.sheetnames:
            wb.create_sheet(sheetname[c])

        wb.save(filename=myFilename)
        wb.close()


# writes weekly related sheetnames to workbook
def write_weekly_sheetname_to_wb():
    sheetname = list(dict.fromkeys(create_sheetnames_weekly()[0]))
    for c in range(len(sheetname)):
        myFilename = path_out + week_xlsx_name

        wb = load_workbook(filename=myFilename)

        if not sheetname in wb.sheetnames:
            wb.create_sheet(sheetname[c])

        wb.save(filename=myFilename)
        wb.close()


# Writes the P values to the spreadsheets
def insert_values_to_spreadsheet():
    sheetname = list(dict.fromkeys(create_sheetnames()[0]))
    date = create_sheetnames()[1]
    p_table = search_for_p()
    myFilename = path_out + month_xlsx_name
    workbook = load_workbook(filename=myFilename)

    for w in range(len(sheetname)):
        ws = workbook[sheetname[w]]

        for n in range(0, 8):
            ws.cell(column=2 + n, row=1, value=pattern[n])

        for item in range(len(date)):
            if date[item][0:7] == sheetname[w]:

                ws.cell(column=1, row=ws.max_row + 1, value=date[item])
                for p in range(0, 8):
                    ws.cell(column=2 + p, row=ws.max_row, value=p_table[item][p])
            workbook.save(filename=myFilename)
        workbook.close()


# inserts weekly values to spreadsheet
def insert_values_to_spreadsheet_weekly():
    sh_name = list(dict.fromkeys(create_sheetnames_weekly()[0]))
    date = create_sheetnames_weekly()[1]
    p_table = search_for_p()

    myFilename = path_out + week_xlsx_name
    workbook = load_workbook(filename=myFilename)

    for w in range(len(sh_name)):
        ws = workbook[sh_name[w]]

        for n in range(0, 8):
            ws.cell(column=2 + n, row=1, value=pattern[n])

        list_week_year = []
        for item in range(len(date)):
            year = int(date[item][:4])
            month = int(date[item][5:7])
            day = int(date[item][8:10])
            week = str(datetime.date(year, month, day).isocalendar()[1])
            list_week_year.append(week + "-" + str(year))

            if list_week_year[item] == sh_name[w]:

                ws.cell(column=1, row=ws.max_row + 1, value=date[item])
                for p in range(0, 8):
                    ws.cell(column=2 + p, row=ws.max_row, value=p_table[item][p])
            workbook.save(filename=myFilename)
        workbook.close()


def insert_monthly_sums_to_spreadsheet():
    sh_name = list(dict.fromkeys(create_sheetnames()[0]))

    myFilename = path_out + month_xlsx_name
    workbook = load_workbook(filename=myFilename)

    for sh in range(len(sh_name)):
        ws = workbook[sh_name[sh]]
        paint_worksheets(ws)

        sums2 = ["=SUM(B1:B24)", "=SUM(C1:C24)", "=SUM(D1:D24)", "=SUM(E1:E24)", "=SUM(F1:F24)", "=SUM(G1:G24)",
                 "=SUM(H1:H24)", "=SUM(I1:I24)", "=SUM(B25:I25)"]
        for s2 in range(len(sums2)):
            ws.cell(column=s2+2, row=25, value=sums2[s2])
        ws.cell(column=10, row=24, value="Monthly Sum")
        keys(ws)


    workbook.save(filename=myFilename)
    workbook.close()


def insert_weekly_sums_to_spreadsheet():
    sh_name = list(dict.fromkeys(create_sheetnames_weekly()[0]))

    myFilename = path_out + week_xlsx_name
    workbook = load_workbook(filename=myFilename)

    for sh in range(len(sh_name)):
        ws = workbook[sh_name[sh]]
        paint_worksheets(ws)

        sums = ["=SUM(B1:B8)", "=SUM(C1:C8)", "=SUM(D1:D8)", "=SUM(E1:E8)", "=SUM(F1:F8)", "=SUM(G1:G8)", "=SUM(H1:H8)",
                "=SUM(I1:I8)", "=SUM(B10:I10)"]

        for sum in range(len(sums)):
            ws.cell(column=sum+2, row=10, value=sums[sum])
        ws.cell(column=10, row=9, value="Weekly Sum")

        keys(ws)

    workbook.save(filename=myFilename)
    workbook.close()


type_of_report = input('What type of report would you like to create? [m - monthly | w - weekly] ').lower()
if type_of_report == 'm':
    check_xls_file()
    search_for_p()
    write_sheetname_to_wb()
    insert_values_to_spreadsheet()
    insert_monthly_sums_to_spreadsheet()

elif type_of_report == 'w':
    week_check_xls_file()
    search_for_p()
    write_weekly_sheetname_to_wb()
    insert_values_to_spreadsheet_weekly()
    insert_weekly_sums_to_spreadsheet()

else:
    print('Only m or w was expected.')
    sys.exit()
